#import dependencies
import re
import openpyxl

#import sheets #TODO: Add user input for sheet selection
workbook_in = openpyxl.load_workbook("01_03_23_FTTT_Seattle & Tacoma.xlsx")
workbook_out = openpyxl.load_workbook("TIPC2.xlsx")
sheet_in = workbook_in.active
sheet_out = workbook_out.active

#initialize array for tercom job numbers
tercomNumArray = [] # I was dumb and initialized this array in the loop and couldn't realize why the script wouldn't append to it for an hour lol

#initialize main loop for each row of sheet_in
for row_in in range(2, sheet_in.max_row):
    #initialize regex strings
    jobNumRegex = re.compile(r'(.\.\d\d\d\d\d\d)')
    inProgressRegex = re.compile(r'In Progress', re.IGNORECASE)
    #regex search for jobs in progress
    inProgress = inProgressRegex.search(str(sheet_in['K{}'.format(row_in)].value))
    #variables to extract raw data from cell
    address = str(sheet_in['N{}'.format(row_in)].value)
    customer = str(sheet_in['B{}'.format(row_in)].value)
    #rexgex search for job number
    #? Maybe better way to do this but script excepts with failed search so exceptions go to none
    try:
        jobNum = jobNumRegex.search(str(sheet_in['G{}'.format(row_in)].value)).group()
    except AttributeError:
        None
    #if job is in progress and has job number initialize loop for sheet out
    #!! I just realized script does not check for existing jobs and will duplicate them. Should be easy to fix.
    if inProgress and jobNum:
        for row_out in range(2, sheet_out.max_row):
            #initialize regex string for tercom job numbers
            #!! Will only find 4 digit tercom job numbers will break in future fix regex before then
            tercomNumRegex = re.compile(r'\d\d\d\d')
            #search for tercom job number and append to array
            #* This code is super subobtimal. Should not be searching the whole sheet for a maximum value each time a row is appended but who cares it works for now
            try:
                tercomNum = tercomNumRegex.search(str(sheet_out['A{}'.format(row_out)].value)).group()
                tercomNumArray.append(tercomNum)
            except AttributeError: #again have to throw failed searches in the trash here
                None
        #insert new row below header
        sheet_out.insert_rows(2)
        #print data in each column of new row
        #?? Could be done better with some sort of array?
        sheet_out.cell(2,1, value=str(int(max(tercomNumArray)) + 1)) #maximum tercom job number incremented by 1
        sheet_out.cell(2,2, value=str(jobNum + " " + address)) #Project name
        sheet_out.cell(2,3, value="Imported from python") #just a bit extra to show what is newly imported #TODO remove or comment out for implementation
        sheet_out.cell(2,5, value=jobNum) #Owner Project Number (N.######)
        sheet_out.cell(2,8, value=address) #address #TODO extract city + state from wire center clli and append to this and their respective columns
        sheet_out.cell(2,9, value="FTTT") #program is only FTTT for this script I think
        sheet_out.cell(2,11, value=customer) #! customer from sheet does not seem like same customer in TIPC review before implement
workbook_out.save("TIPC2UPD.xlsx") #Save to new sheet
print("Success!!!") #validate successful script end in console
